from __future__ import annotations

import asyncio
import hashlib
import io
import json
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from uuid import uuid4

import sqlalchemy as sa
from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.infrastructure.database.models import HSCode, TariffMeasure


def _norm(s: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else " " for ch in s).strip()


def _find_col(headers: list[str], *, must: list[str], any_of: list[str] | None = None) -> int | None:
    any_of = any_of or []
    for i, h in enumerate(headers):
        nh = _norm(h)
        ok = all(token in nh for token in must)
        if not ok:
            continue
        if any_of and not any(token in nh for token in any_of):
            continue
        return i
    return None


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    return str(v).strip() or None


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _to_str(v)
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except Exception:
        pass
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue
    return None


def _to_decimal(v: Any) -> Decimal | None:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = _to_str(v)
    if not s:
        return None
    s = s.replace(",", ".")
    s = s.replace("%", "").strip()
    try:
        return Decimal(s)
    except Exception:
        return None


def _digits(v: str | None) -> str | None:
    if not v:
        return None
    d = "".join(ch for ch in v if ch.isdigit())
    return d or None


def _parse_duty_cell(duty_val: Any) -> tuple[Decimal | None, Decimal | None, str | None]:
    if duty_val is None:
        return None, None, None
    if isinstance(duty_val, str):
        s = duty_val.strip()
        if not s:
            return None, None, None
        if "%" in s:
            return _to_decimal(s), None, None

        m = re.search(r"([0-9]+(?:[.,][0-9]+)?)\s*([A-Z]{3})\s*/\s*([A-Z0-9]{1,10})", s)
        if m:
            amount = _to_decimal(m.group(1))
            unit = f"{m.group(2)}/{m.group(3)}"
            return None, amount, unit if len(unit) <= 50 else None

        m = re.search(r"([0-9]+(?:[.,][0-9]+)?)\s*([A-Z]{3})\s*([A-Z0-9]{1,10})", s)
        if m:
            amount = _to_decimal(m.group(1))
            unit = f"{m.group(2)}/{m.group(3)}"
            return None, amount, unit if len(unit) <= 50 else None

        return _to_decimal(s), None, None

    if isinstance(duty_val, (int, float, Decimal)):
        return Decimal(str(duty_val)), None, None
    return None, None, None


def inspect_xlsx_bytes(data: bytes, *, max_rows: int = 5) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: dict[str, Any] = {"sheets": []}
    for name in wb.sheetnames:
        ws = wb[name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        headers = [str(h).strip() if h is not None else "" for h in (header_row or [])]
        samples: list[list[Any]] = []
        for _ in range(max_rows):
            r = next(rows_iter, None)
            if r is None:
                break
            samples.append(list(r))
        out["sheets"].append({"name": name, "headers": headers, "sample_rows": samples})
    return out


async def inspect_xlsx(data: bytes, *, max_rows: int = 5) -> dict[str, Any]:
    return await asyncio.to_thread(inspect_xlsx_bytes, data, max_rows=max_rows)


async def ingest_nomenclature_en_xlsx(db: AsyncSession, data: bytes) -> dict[str, Any]:
    payload = await inspect_xlsx(data, max_rows=1)
    sheets = payload.get("sheets") or []
    if not sheets:
        raise ValueError("No sheets found")
    sheet = sheets[0]
    headers: list[str] = list(sheet.get("headers") or [])

    code_col = _find_col(headers, must=["code"])
    if code_col is None:
        code_col = _find_col(headers, must=["goods"], any_of=["code", "nomenclature"])
    desc_col = _find_col(headers, must=["description"])
    if desc_col is None:
        desc_col = _find_col(headers, must=["desc"])

    if code_col is None or desc_col is None:
        raise ValueError(f"Could not detect columns. Headers={headers}")

    wb = await asyncio.to_thread(load_workbook, io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)

    upserted = 0
    for row in rows:
        code_raw = _to_str(row[code_col] if code_col < len(row) else None)
        hs = _digits(code_raw)
        if not hs or len(hs) < 6:
            continue
        desc = _to_str(row[desc_col] if desc_col < len(row) else None) or f"HS {hs}"
        stmt = (
            insert(HSCode)
            .values(
                code=hs,
                jurisdiction="EU",
                description=desc,
                parent_code=None,
                level=len(hs),
                supplementary_unit=None,
                valid_from=date.today(),
                valid_to=None,
            )
            .on_conflict_do_update(
                index_elements=[HSCode.code],
                set_={"jurisdiction": "EU", "description": desc, "level": len(hs), "valid_to": None},
            )
        )
        await db.execute(stmt)
        upserted += 1
        if upserted % 500 == 0:
            await db.commit()

    await db.commit()
    return {"hs_codes_upserted": upserted}


def _measure_type_from_row(mt_id: str | None, mt_name: str | None) -> str | None:
    if mt_id:
        mt_id = mt_id.strip()
        mapping = {
            "103": "MFN",
            "142": "PREFERENTIAL",
            "551": "ANTI_DUMPING",
            "552": "COUNTERVAILING",
            "112": "SUSPENSION",
            "115": "END_USE",
        }
        if mt_id in mapping:
            return mapping[mt_id]
    if mt_name:
        t = mt_name.lower()
        if "third" in t and "duty" in t:
            return "MFN"
        if "anti" in t and "dump" in t:
            return "ANTI_DUMPING"
        if "counter" in t and "vail" in t:
            return "COUNTERVAILING"
        if "suspension" in t:
            return "SUSPENSION"
        if "preference" in t or "preferential" in t:
            return "PREFERENTIAL"
    return None


async def ingest_duties_import_xlsx(db: AsyncSession, data: bytes) -> dict[str, Any]:
    meta = await inspect_xlsx(data, max_rows=1)
    sheets = meta.get("sheets") or []
    if not sheets:
        raise ValueError("No sheets found")
    headers: list[str] = list(sheets[0].get("headers") or [])

    code_col = _find_col(headers, must=["goods", "code"]) or _find_col(headers, must=["code"])
    origin_code_col = _find_col(headers, must=["origin", "code"])
    origin_name_col = _find_col(headers, must=["origin"])
    mt_code_col = _find_col(headers, must=["meas", "type", "code"]) or _find_col(headers, must=["measure", "type", "code"])
    mt_name_col = _find_col(headers, must=["measure", "type"])
    duty_col = _find_col(headers, must=["duty"])
    legal_base_col = _find_col(headers, must=["legal", "base"])
    start_col = _find_col(headers, must=["start", "date"])
    end_col = _find_col(headers, must=["end", "date"])

    if code_col is None or duty_col is None:
        raise ValueError(f"Could not detect required columns. Headers={headers}")

    wb = await asyncio.to_thread(load_workbook, io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)

    measures_upserted = 0
    rows_seen = 0
    hs_codes_upserted = 0
    seen_hs: set[str] = set()

    for row in rows:
        rows_seen += 1
        code_raw = _to_str(row[code_col] if code_col < len(row) else None)
        hs = _digits(code_raw)
        if not hs or len(hs) < 6:
            continue

        if hs not in seen_hs:
            stmt = (
                insert(HSCode)
                .values(
                    code=hs,
                    jurisdiction="EU",
                    description=f"HS {hs}",
                    parent_code=None,
                    level=len(hs),
                    supplementary_unit=None,
                    valid_from=date.today(),
                    valid_to=None,
                )
                .on_conflict_do_update(
                    index_elements=[HSCode.code],
                    set_={"jurisdiction": "EU", "level": len(hs), "valid_to": None},
                )
            )
            await db.execute(stmt)
            hs_codes_upserted += 1
            seen_hs.add(hs)

        origin = None
        origin_code = _to_str(row[origin_code_col] if origin_code_col is not None and origin_code_col < len(row) else None)
        if origin_code:
            o = origin_code.strip().upper()
            if o not in {"1011", "ERGA OMNES", "ALL", "WORLD"}:
                origin = o[:5]
        else:
            origin_raw = _to_str(row[origin_name_col] if origin_name_col is not None and origin_name_col < len(row) else None)
            if origin_raw:
                o = origin_raw.strip().upper()
                if o not in {"1011", "ERGA OMNES", "ALL", "WORLD"}:
                    origin = o[:5]

        mt_id = _to_str(row[mt_code_col] if mt_code_col is not None and mt_code_col < len(row) else None)
        mt_name = _to_str(row[mt_name_col] if mt_name_col is not None and mt_name_col < len(row) else None)
        measure_type = _measure_type_from_row(mt_id, mt_name)
        if not measure_type:
            continue

        duty_val = row[duty_col] if duty_col < len(row) else None
        rate = _to_decimal(duty_val)
        valid_from = _to_date(row[start_col] if start_col is not None and start_col < len(row) else None) or date.today()
        valid_to = _to_date(row[end_col] if end_col is not None and end_col < len(row) else None)

        legal_base = _to_str(row[legal_base_col] if legal_base_col is not None and legal_base_col < len(row) else None)

        raw_row = {headers[i] if i < len(headers) else f"col_{i}": _to_str(v) for i, v in enumerate(row)}
        raw_row = {k: v for k, v in raw_row.items() if k and v is not None}
        raw_digest = hashlib.sha256(json.dumps(raw_row, sort_keys=True).encode("utf-8")).hexdigest()
        source_measure_id = f"TARIC_XLSX:{raw_digest}"

        rate_ad_valorem: Decimal | None
        rate_specific_amount: Decimal | None
        rate_specific_unit: str | None
        rate_ad_valorem, rate_specific_amount, rate_specific_unit = _parse_duty_cell(duty_val)
        if rate_ad_valorem is None and rate_specific_amount is None:
            rate_ad_valorem = rate
        if rate_specific_unit and len(rate_specific_unit) > 50:
            rate_specific_unit = None

        stmt = (
            insert(TariffMeasure)
            .values(
                id=uuid4(),
                hs_code=hs,
                jurisdiction="EU",
                measure_type=measure_type,
                country_of_origin=origin,
                preferential_agreement=None,
                rate_ad_valorem=rate_ad_valorem,
                rate_specific_amount=rate_specific_amount,
                rate_specific_unit=rate_specific_unit,
                rate_minimum=None,
                rate_maximum=None,
                agricultural_component=None,
                quota_id=None,
                suspension=measure_type in {"SUSPENSION", "END_USE"},
                measure_condition=None,
                raw_json=raw_row,
                valid_from=valid_from,
                valid_to=valid_to,
                source_dataset="TARIC",
                source_measure_id=source_measure_id,
                ingested_at=datetime.utcnow(),
            )
            .on_conflict_do_update(
                index_elements=[TariffMeasure.source_dataset, TariffMeasure.source_measure_id],
                index_where=sa.text("source_measure_id IS NOT NULL"),
                set_={
                    "hs_code": hs,
                    "measure_type": measure_type,
                    "country_of_origin": origin,
                    "rate_ad_valorem": rate_ad_valorem,
                    "rate_specific_amount": rate_specific_amount,
                    "rate_specific_unit": rate_specific_unit,
                    "suspension": measure_type in {"SUSPENSION", "END_USE"},
                    "raw_json": raw_row,
                    "valid_from": valid_from,
                    "valid_to": valid_to,
                    "ingested_at": datetime.utcnow(),
                },
            )
        )
        await db.execute(stmt)
        measures_upserted += 1
        if measures_upserted % 500 == 0:
            await db.commit()

    await db.commit()
    return {"rows_seen": rows_seen, "hs_codes_upserted": hs_codes_upserted, "measures_upserted": measures_upserted}
